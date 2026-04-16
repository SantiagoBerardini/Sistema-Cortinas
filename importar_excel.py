import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook

# =======================
# CONFIGURACIÓN
# =======================

EXCEL_FILE = 'Lista_Roller_Calculo_automatico_1450_02_03_2026__1_.xlsx'
OUTPUT_JSON = 'productos_importar.json'

# Mapeo de tipos de tela → factor desperdicio
FACTORES_DESPERDICIO = {
    'VOILE': 1.0,  # Sin desperdicio
    'BLACK OUT VINILICO SOSTEN MUTUO': 1.1,
    'BLACK OUT VINILICO CADAC LISO': 1.1,
    'Black out poliester Goma': 1.1,
    'TSHA': 1.1,
    'VENECIA': 1.2,  # Más variación de colores
    'VENECIA ANCHA': 1.15,
    'HINDU': 1.2,
    'GAZA': 1.1,
    'RAMIO (RUSTICA)': 1.1,
    'SCREEN': 1.0,  # Sin desperdicio (screens)
    'SCREEN BICOLOR': 1.0,
    'SCREEN MILAN': 1.0,
    'Traslucida DEVON': 1.0,  # Sin desperdicio
}

# =======================
# FUNCIONES
# =======================

def normalizar_nombre(nombre):
    """Limpia y normaliza nombres de telas"""
    if not nombre:
        return None
    return nombre.strip().title()

def obtener_factor_desperdicio(nombre_tela):
    """Busca el factor desperdicio más cercano"""
    if not nombre_tela:
        return 1.1
    
    nombre_upper = nombre_tela.upper()
    
    for tipo, factor in FACTORES_DESPERDICIO.items():
        if tipo.upper() in nombre_upper:
            return factor
    
    # Default si no encontró coincidencia
    if 'SCREEN' in nombre_upper or 'VOILE' in nombre_upper:
        return 1.0
    elif 'VENECIA' in nombre_upper:
        return 1.15
    else:
        return 1.1

def limpiar_precio(precio):
    """Convierte precio a float, maneja '-' y strings vacíos"""
    if precio is None or precio == '' or precio == '-' or precio == '----------' or precio == '----------.':
        return None
    try:
        return float(precio)
    except:
        return None

def procesar_excel():
    """Lee Excel y genera JSON de productos para importar"""
    
    print("📂 Leyendo Excel...")
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Leer también con pandas para datos limpios
    df = pd.read_excel(EXCEL_FILE, sheet_name=0)
    
    productos = []
    contador = 0
    
    # Procesar filas (saltando header en fila 1)
    for row in range(2, 113):  # Filas 2 a 112
        try:
            # Obtener valores de columnas principales
            tela = ws[f'A{row}'].value
            ancho = ws[f'B{row}'].value
            colores = ws[f'C{row}'].value
            precio_roller = ws[f'F{row}'].value
            precio_bandas = ws[f'G{row}'].value
            
            # Saltar si no hay nombre de tela
            if not tela:
                continue
            
            # Limpiar datos
            tela = normalizar_nombre(tela)
            ancho = float(ancho) if ancho else None
            precio_roller = limpiar_precio(precio_roller)
            precio_bandas = limpiar_precio(precio_bandas)
            
            # Saltar si no hay ancho
            if not ancho:
                continue
            
            # Obtener factor desperdicio
            factor = obtener_factor_desperdicio(tela)
            
            # Sin precio no es útil (es una tela que no venden)
            if not precio_roller:
                print(f"  ⚠️  Fila {row}: {tela} {ancho}m - SIN PRECIO (saltando)")
                continue
            
            # Crear nombre completo del producto
            nombre_producto = f"{tela} {ancho}m"
            
            # Calcular precio por m² (aproximado)
            # Asumir que el precio es para un metro lineal de 1m de alto
            # Entonces: precio_m2 = precio / 1
            precio_m2 = precio_roller / 1  # Ajustar si es necesario
            
            producto = {
                'nombre': nombre_producto,
                'sku': f"{tela.upper().replace(' ', '')}_{ancho}m".replace('.', ''),
                'descripcion': f"Tela {tela} - Ancho: {ancho}m - Colores: {colores or 'Varios'}",
                'categoria_id': 1,  # Cambiar según tu ID de categoría TELAS
                'tipo_medida': 'metros_cuadrados',
                'ancho_rollo': ancho,
                'largo_rollo': 30,
                'factor_desperdicio': factor,
                'stock_actual': 0,  # IMPORTANTE: Usuario debe actualizar manualmente
                'stock_minimo': 50,  # Default: 50 m²
                'costo_unitario': precio_m2,
                'es_producto_base': True,
                'colores': colores or 'Varios',
                'precio_roller_original': float(precio_roller),
                'precio_bandas_original': float(precio_bandas) if precio_bandas else None
            }
            
            productos.append(producto)
            contador += 1
            print(f"  ✅ Fila {row}: {nombre_producto} - Factor: {factor}x - ${precio_roller:.2f}")
        
        except Exception as e:
            print(f"  ❌ Error en fila {row}: {str(e)}")
            continue
    
    # Guardar a JSON
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump({
            'fecha_exportacion': datetime.now().isoformat(),
            'total_productos': len(productos),
            'productos': productos
        }, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Exportación completa!")
    print(f"   Total productos procesados: {contador}")
    print(f"   Archivo: {OUTPUT_JSON}")
    return productos

def generar_sql(productos):
    """Genera script SQL para insertarlos en Supabase"""
    
    sql_file = 'insertar_productos.sql'
    
    with open(sql_file, 'w', encoding='utf-8') as f:
        f.write("-- Script generado automáticamente\n")
        f.write(f"-- Fecha: {datetime.now().isoformat()}\n")
        f.write(f"-- Total de inserts: {len(productos)}\n\n")
        f.write("-- EJECUTAR EN: Supabase SQL Editor\n")
        f.write("-- IMPORTANTE: Cambiar categoria_id según tu base de datos\n\n")
        
        for i, prod in enumerate(productos, 1):
            sql = f"""
INSERT INTO productos (
    categoria_id,
    nombre,
    sku,
    descripcion,
    tipo_medida,
    ancho_rollo,
    largo_rollo,
    factor_desperdicio,
    stock_actual,
    stock_minimo,
    costo_unitario,
    es_producto_base,
    activo
) VALUES (
    {prod['categoria_id']},
    '{prod['nombre'].replace("'", "''")}',
    '{prod['sku'].replace("'", "''")}',
    '{prod['descripcion'].replace("'", "''")}',
    '{prod['tipo_medida']}',
    {prod['ancho_rollo']},
    {prod['largo_rollo']},
    {prod['factor_desperdicio']},
    {prod['stock_actual']},
    {prod['stock_minimo']},
    {prod['costo_unitario']},
    {str(prod['es_producto_base']).lower()},
    true
);
"""
            f.write(sql)
            if i % 10 == 0:
                f.write(f"\n-- {i} de {len(productos)} inserts\n\n")
    
    print(f"\n📝 SQL generado: {sql_file}")
    print(f"   {len(productos)} sentencias INSERT listas")

def generar_reporte(productos):
    """Genera reporte en texto"""
    
    report_file = 'reporte_importacion.txt'
    
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("REPORTE DE IMPORTACIÓN - LISTA DE ROLLERS\n")
        f.write("=" * 80 + "\n\n")
        
        f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total de productos: {len(productos)}\n")
        f.write(f"Archivo Excel: {EXCEL_FILE}\n\n")
        
        f.write("PRODUCTOS IMPORTADOS:\n")
        f.write("-" * 80 + "\n")
        f.write(f"{'#':<4} {'NOMBRE':<35} {'FACTOR':<8} {'PRECIO':<12}\n")
        f.write("-" * 80 + "\n")
        
        for i, prod in enumerate(productos, 1):
            f.write(f"{i:<4} {prod['nombre']:<35} {prod['factor_desperdicio']:<8.2f} "
                   f"${prod['costo_unitario']:<11.2f}\n")
        
        f.write("\n" + "=" * 80 + "\n")
        f.write("RESUMEN POR FACTOR DESPERDICIO:\n")
        f.write("-" * 80 + "\n")
        
        # Agrupar por factor
        por_factor = {}
        for prod in productos:
            factor = prod['factor_desperdicio']
            if factor not in por_factor:
                por_factor[factor] = []
            por_factor[factor].append(prod['nombre'])
        
        for factor in sorted(por_factor.keys()):
            productos_con_factor = por_factor[factor]
            f.write(f"\nFactor {factor}x ({(factor-1)*100:.0f}% desperdicio): "
                   f"{len(productos_con_factor)} productos\n")
            for nombre in productos_con_factor[:5]:  # Mostrar primeros 5
                f.write(f"  • {nombre}\n")
            if len(productos_con_factor) > 5:
                f.write(f"  ... y {len(productos_con_factor)-5} más\n")
        
        f.write("\n" + "=" * 80 + "\n")
        f.write("INSTRUCCIONES PARA IMPORTAR:\n")
        f.write("-" * 80 + "\n")
        f.write("""
1. VERIFICAR DATOS EN JSON:
   - Abrir: productos_importar.json
   - Revisar factores desperdicio
   - Ajustar stock_minimo si es necesario
   - Cambiar categoria_id si es diferente

2. GENERAR SQL:
   - Abrir: insertar_productos.sql
   - IMPORTANTE: Cambiar "categoria_id": 1 si tus IDs son diferentes
   - Copiar el contenido

3. EJECUTAR EN SUPABASE:
   - Ir a: https://app.supabase.com
   - Proyecto → SQL Editor
   - Nuevo query
   - Pegar el SQL
   - Ejecutar

4. VALIDAR:
   - Ver productos en tabla "productos"
   - Confirmar cantidad de filas = """ + str(len(productos)) + """
   - Revisar un producto para confirmar estructura

5. ACTUALIZAR STOCKS:
   - MANUALMENTE llevar stock_actual a valores reales
   - Los productos importados llegan con 0 m² de stock
   - Esto es INTENCIONAL para evitar errores
""")
    
    print(f"\n📊 Reporte generado: {report_file}")

# =======================
# MAIN
# =======================

if __name__ == '__main__':
    print("\n" + "=" * 80)
    print("🔄 IMPORTADOR DE EXCEL - BM CORTINAS")
    print("=" * 80 + "\n")
    
    # Procesar Excel
    productos = procesar_excel()
    
    if not productos:
        print("❌ No se encontraron productos para importar")
        exit(1)
    
    # Generar archivos
    generar_sql(productos)
    generar_reporte(productos)
    
    print("\n" + "=" * 80)
    print("✅ ARCHIVOS GENERADOS:")
    print(f"   1. {OUTPUT_JSON} - Datos en formato JSON")
    print(f"   2. insertar_productos.sql - Script SQL para ejecutar")
    print(f"   3. reporte_importacion.txt - Reporte detallado")
    print("=" * 80 + "\n")
    
    print("📋 PRÓXIMOS PASOS:")
    print("   1. Revisar reporte_importacion.txt")
    print("   2. Editar productos_importar.json si necesario")
    print("   3. Copiar SQL de insertar_productos.sql")
    print("   4. Ejecutar en Supabase SQL Editor")
    print("   5. Actualizar stocks manualmente\n")